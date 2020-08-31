import openpyxl


"""
"id","parent_id:id","code","name","tax_classification_id:id"
"""


def load():
    book = openpyxl.load_workbook("QM产线清单.xlsx")
    sheet = book["QM税收编码整理"]

    iter = sheet.iter_rows()
    next(iter)
    cates = {}
    tcs = {}
    seqs = {}
    for row in iter:
        values = [x.value for x in row]
        values = map(lambda x: "" if not x else str(x).strip(), values)

        code, lvl1_name, lvl1_code, lvl2_name, lvl2_code, lvl3_name, lvl3_code, _, tc_code, tc18_code, tc_name = tuple(
            values
        )

        if tc_code and tc_code.strip() and tc_code not in tcs:
            tcs[tc_code] = {
                "id": f"tax_classification_{tc_code}",
                "name": tc_name,
                "code": tc_code,
                "code18": tc18_code,
            }

        lvl3_full_code = f"{lvl1_code}{lvl2_code}{lvl3_code}"
        lvl2_full_code = f"{lvl1_code}{lvl2_code}"

        if lvl3_name and lvl3_code:
            cates[lvl3_full_code] = {
                "id": f"product_category_{lvl3_full_code}",
                "code": lvl3_code,
                "name": lvl3_name,
                "parent_id": f"product_category_{lvl2_full_code}",
                "tax_classification_id": f"tax_classification_{tc_code}"
                if tc_code and tc_name
                else "",
                "level": 3,
            }

            if lvl3_code not in seqs:
                seqs[lvl3_full_code] = {
                    "id": f"seq_product_{lvl3_full_code}",
                    "name": f"Product {lvl3_full_code}",
                    "code": f"product.{lvl3_full_code}",
                    "padding": 4,
                }

        if lvl2_full_code not in cates:
            cates[lvl2_full_code] = {
                "id": f"product_category_{lvl2_full_code}",
                "code": lvl2_code,
                "name": lvl2_name,
                "parent_id": f"product_category_{lvl1_code}",
                "tax_classification_id": "",
                "level": 2,
            }

        if lvl1_code not in cates:
            cates[lvl1_code] = {
                "id": f"product_category_{lvl1_code}",
                "code": lvl1_code,
                "name": lvl1_name,
                "level": 1,
            }

    cate_sheet = book.create_sheet("类别")
    tc_sheet = book.create_sheet("税收分类")
    seq_sheet = book.create_sheet("序号")

    for c in sorted(cates.values(), key=lambda x: x["level"]):
        cate_sheet.append(
            (
                c["id"],
                c.get("parent_id") or "",
                c["code"],
                c["name"],
                c.get("tax_classification_id") or "",
            )
        )

    for t in sorted(tcs.values(), key=lambda x: x["code"]):
        tc_sheet.append((t["id"], t["code"], t["code18"], t["name"]))

    for seq in sorted(seqs.values(), key=lambda x: x["code"]):
        seq_sheet.append((seq["id"], seq["code"], seq["name"], 4))

    book.save("qm.xlsx")


if __name__ == "__main__":
    load()
